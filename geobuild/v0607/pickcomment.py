#!/bin/env python 
# 
#  Extract geometry parameter from geobuild.py comment field
#

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.alignment import Alignment


# ===========================================================
if __name__ == "__main__":

    execfile("version.py")
    version = _VERSION if _VERSION[0:1] != "-" else _VERSION[1:]

    geofile="geobuild.py"

    comments = []
    category = ""
    for line in open(geofile):
        if "#" not in line or line[0:1] == "#" or "###" in line:
            continue
        # print line[:-1]
        if "#@" in line:
            (dummy, category) = line[:-1].split("@",1)
            category=category.strip()
            continue
        (varfield, comment) = line[:-1].split("#",1)
        var = varfield.strip()
        print category+" | "+ var + " | "+comment
        if len(var) == 0:
            comments[-1][3] += comment
        elif ":" in var:
            var1 = var.replace("\"","").replace("{","").replace("}","")
            (vname, value) = var1.split(":")
            value=value.replace(",","").strip()
            comments.append([category,vname.strip(),value,comment.strip()])

    # Output to excel book

    # for com in comments:
    #    print com

    excelbook="geoparam.xlsx"
    wb = Workbook()
    ws = wb.create_sheet(title=version)
    
    font = Font(name="Arial", size=12)

    maxchar = [0,0,0,0]
    header = ["Category", "Varname", "value(cm)", "Meaning"]
    for ic in range(0, len(header)):
        ws.cell(row=1, column=ic+1).value=header[ic]
        ws.cell(row=1, column=ic+1).font=font
        maxchar[ic] = len(header[ic])

    for ir in range(0, len(comments)):
        for ic in range(0, len(header)):
            ws.cell(row=ir+2, column=ic+1).value = comments[ir][ic]
            ws.cell(row=ir+2, column=ic+1).font = font
            if len(comments[ir][ic]) > maxchar[ic]:
                maxchar[ic] = len(comments[ir][ic])
            if ic == 2:
                ws.cell(row=1, column=ic+1).alignment = Alignment(horizontal="right")

    colname=["A","B","C","D"]
    for ic in range(0, len(header)):
        ws.column_dimensions[colname[ic]].width = int(maxchar[ic]*1.4)

    wb.save(excelbook)

    print "Geometry parameters are written in "+excelbook
